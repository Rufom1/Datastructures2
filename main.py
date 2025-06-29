# Author: Michael Rufo 

from package import Package
from hashtable import HashTable
from openpyxl import load_workbook
from pathlib import Path
from enum import Enum

class DeliveryStatus(Enum):
    NOT_STARTED = "Not Started"
    IN_TRANSIT = "In Transit"
    DELIVERED = "Delivered"
    DELAYED = "Delayed"

def lookup_package(package_id, package_table):
    """
    Function to look up a package by its ID in the package list.
    :param package_id: The ID of the package to look up.
    :param package_list: The table of packages where each package is stored with its ID as the key.
    :return: The package if found, otherwise None.
    """

    if package_id in package_table:
        return package_table[package_id].lookupPackage()
    else:
        return None



def main():
    '''
    Main function to deliver packages using a self-adjusting algorithm and a hash table.

    Initialize a hash table to store package data
    Initialize a hash table to store delivery data
    Initialize a set of trucks to deliver packages 
        contaning truck objects with attributes such as truck ID, and packages assigned, distance traveled, package weight
            and methods to update packages, deliver packages, calculate distance, and update status

    read package and delivery data from file 
    for each package in the package data:
        assign a unique ID to the package
        assign a delivery address, deadline, city, zip code, weight, and delivery status
        add the package to the package table

        If the package has a duplicate address, set the duplicateAddress attribute to True
        If the package has a duplicate address, incriment the duplicateAddressPointer 
            and add it to the list of packages with the same address 
        calculate the distance from the hub to the package address      
        assign the nearest neighbor ID and distance to the package

    Intialize a queue to manage loading packaged onto trucks
        containing package objects

    Sort packages into queue based on distance from hub and delivery deadline

    Sort packages into trucks based on the nearest neighbor algorithm
        for each truck:
            while there are packages in the queue:
                if truck is empty:
                    load the first package from the queue onto the truck
                else:
                    check if the truck can carry the package based on its weight and capacity
                    if the truck can carry the package:
                        check if duplicate address exists
                        if duplicate address exists:
                            skip to duplicate address pointer
                            load duplicates onto truck until truck is full or no more duplicates
                            decrement the package iterator back to skipped package
                        else
                            load the package onto the truck
                    else:
                        dequeue the next truck and repeat the process
    '''

    parent_dir = Path(__file__).resolve().parent.parent

    wgu_packages_file = load_workbook(f'{parent_dir}/WGUPS Package File.xlsx')
    package_distances_file = load_workbook(f'{parent_dir}/WGUPS Distance Table.xlsx')

    delivery_array = [row for row in wgu_packages_file.active.iter_rows(min_row=8, max_row=48, min_col=1, max_col=8, values_only=True)]
    distance_array = [row for row in package_distances_file.active.iter_rows(min_row=8, max_row=35, min_col=1, max_col=29, values_only=True)]

    deliveriesTable = HashTable()
    packageTable = HashTable()

    for i in range(1, len(delivery_array)):
        row = delivery_array[i]
        package_id = row[0]
        address = str(row[1])
        deadline = row[5]
        city = row[2]
        zip_code = row[4]
        weight = row[6]
        deadline = row[5]
        notes = row[6]
        delivery_status = DeliveryStatus.NOT_STARTED.value

        package = Package(package_id, address, deadline, city, zip_code, weight, delivery_status)

        if not deliveriesTable.exists(address):
            deliveriesTable.insert(address, [package])
        else:
            existing_packages = deliveriesTable.get(address)
            existing_packages.append(package)
            deliveriesTable.insert(address, existing_packages)

    for i in range(1, len(distance_array)):
        pass
    
if __name__ == "__main__":
    main()
    # Example usage of lookup_package
    # package_table = HashTable()
    # package_table.insert(1, Package(1, "123 Main St", "10:30 AM", "Salt Lake City", "84101", 5, "Delivered"))
    # package = lookup_package(1, package_table)
    # if package:
    #     print(f"Package found: {package}")
    # else:
    #     print("Package not found.")